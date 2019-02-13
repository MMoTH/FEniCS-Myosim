import numpy as np
import pandas as pd
from openpyxl import Workbook
from lxml import etree
from scipy import signal
#from scipy.integrate import solve_ivp
from scipy import integrate 
from scipy.constants import mmHg as mmHg_in_pascals

try:
    import Python_MyoSim.half_sarcomere.half_sarcomere as hs
except:
    import sys
    sys.path.append('c:\\ken\\github\\campbellmusclelab\\python\\modules')
    import Python_MyoSim.half_sarcomere.half_sarcomere as hs

class single_circulation():
    """Class for a single ventricle circulation"""

    from .display import display_simulation, display_flows, display_pv_loop
#    from .driver import return_sim_struct_from_xml_file, \
#        run_simulation_from_xml_file

    def __init__(self, single_circulation_simulation, xml_file_string = None):
        # Pull off stuff
        self.input_xml_file_string = xml_file_string
        
        self.simulation_parameters = \
            single_circulation_simulation.simulation_parameters

        self.output_parameters = \
            single_circulation_simulation.output_parameters

        self.output_buffer_size = \
            int(single_circulation_simulation.simulation_parameters.
                no_of_time_points.cdata)

        # Look for perturbations
        self.volume_perturbation = np.zeros(self.output_buffer_size+1)
        if (hasattr(single_circulation_simulation, 'perturbations')):
            if hasattr(single_circulation_simulation.perturbations, 'volume'):
                temp = single_circulation_simulation.perturbations.volume
                start_index = int(temp.start_index.cdata)
                stop_index = int(temp.stop_index.cdata)
                increment = float(temp.increment.cdata)
                self.volume_perturbation[(start_index+1):(stop_index+1)] =\
                    increment

        # Initialize circulation object using data from the sim_object
        circ_params = single_circulation_simulation.circulation

        self.no_of_compartments = int(circ_params.no_of_compartments.cdata)
        self.blood_volume = float(circ_params.blood.volume.cdata)

        self.aorta_resistance = float(circ_params.aorta.resistance.cdata)
        self.aorta_compliance = float(circ_params.aorta.compliance.cdata)

        self.arteries_resistance = float(circ_params.arteries.resistance.cdata)
        self.arteries_compliance = float(circ_params.arteries.compliance.cdata)

        self.arterioles_resistance = float(circ_params.arterioles.resistance.cdata)
        self.arterioles_compliance = float(circ_params.arterioles.compliance.cdata)

        self.capillaries_resistance = float(circ_params.capillaries.resistance.cdata)
        self.capillaries_compliance = float(circ_params.capillaries.compliance.cdata)

        self.veins_resistance = float(circ_params.veins.resistance.cdata)
        self.veins_compliance = float(circ_params.veins.compliance.cdata)

        self.ventricle_resistance = \
            float(circ_params.ventricle.resistance.cdata)
        self.ventricle_wall_volume = \
            float(circ_params.ventricle.wall_volume.cdata)
        self.ventricle_slack_volume = \
            float(circ_params.ventricle.slack_volume.cdata)

        # Initialise the resistance and compliance arrays for calcuations
        self.resistance = np.array([self.aorta_resistance,
                                    self.arteries_resistance,
                                    self.arterioles_resistance,
                                    self.capillaries_resistance,
                                    self.veins_resistance,
                                    self.ventricle_resistance])
        self.compliance = np.array([self.aorta_compliance,
                                    self.arteries_compliance,
                                    self.arterioles_compliance,
                                    self.capillaries_compliance,
                                    self.veins_compliance,
                                    0])

        # Pull off the half_sarcomere parameters
        hs_params = single_circulation_simulation.half_sarcomere
        self.hs = hs.half_sarcomere(hs_params, self.output_buffer_size)

        # Deduce the hsl where force is zero and set the hsl to that length
        slack_hsl = self.hs.myof.return_hs_length_for_force(0.0)
        self.hs.update_simulation(0.0,(slack_hsl - self.hs.hs_length), 0.0)

        # Deduce the slack circumference of the ventricle and set that
        self.lv_circumference = \
            self.return_lv_circumference(self.ventricle_slack_volume)
        
        print("hsl: %f" % self.hs.hs_length)
        print("slack hsl: %f" % slack_hsl)
        print("slack_lv_circumference %f" % self.lv_circumference)


        # Set the initial volumes with most of the blood in the veins
        initial_ventricular_volume = 1.5 * self.ventricle_slack_volume
        self.v = np.zeros(self.no_of_compartments)
        self.v[-2] = self.blood_volume - initial_ventricular_volume
        self.v[-1] = initial_ventricular_volume

        # Deduce the pressures
        self.p = np.zeros(self.no_of_compartments)
        for i in np.arange(0, self.no_of_compartments-1):
            self.p[i] = self.v[i] / self.compliance[i]
        self.p[-1] = self.return_lv_pressure(self.v[-1])

        # Create a pandas data structure to store data
        self.sim_time = 0.0
        self.data_buffer_index = 0
        self.data = pd.DataFrame({'time': np.zeros(self.output_buffer_size),
                                  'pressure_aorta':
                                      np.zeros(self.output_buffer_size),
                                  'pressure_arteries':
                                      np.zeros(self.output_buffer_size),
                                  'pressure_arterioles':
                                      np.zeros(self.output_buffer_size),
                                  'pressure_capillaries':
                                      np.zeros(self.output_buffer_size),
                                  'pressure_veins':
                                      np.zeros(self.output_buffer_size),
                                  'pressure_ventricle':
                                      np.zeros(self.output_buffer_size),
                                  'volume_aorta':
                                      np.zeros(self.output_buffer_size),
                                  'volume_arteries':
                                      np.zeros(self.output_buffer_size),
                                  'volume_arterioles':
                                      np.zeros(self.output_buffer_size),
                                  'volume_capillaries':
                                      np.zeros(self.output_buffer_size),
                                  'volume_veins':
                                      np.zeros(self.output_buffer_size),
                                  'volume_ventricle':
                                      np.zeros(self.output_buffer_size),
                                  'flow_ventricle_to_aorta':
                                      np.zeros(self.output_buffer_size),
                                  'flow_aorta_to_arteries':
                                      np.zeros(self.output_buffer_size),
                                  'flow_arteries_to_arterioles':
                                      np.zeros(self.output_buffer_size),
                                  'flow_arterioles_to_capillaries':
                                      np.zeros(self.output_buffer_size),
                                  'flow_capillaries_to_veins':
                                      np.zeros(self.output_buffer_size),
                                  'flow_veins_to_ventricle':
                                      np.zeros(self.output_buffer_size),
                                  'volume_perturbation':
                                      np.zeros(self.output_buffer_size),
                                  'ventricle_wall_volume':
                                      np.zeros(self.output_buffer_size)})
        # Store the first values
        self.data.at[0, 'pressure_aorta'] = self.p[0]
        self.data.at[0, 'pressure_arteries'] = self.p[1]
        self.data.at[0, 'pressure_arterioles'] = self.p[2]
        self.data.at[0, 'pressure_capillaries'] = self.p[3]
        self.data.at[0, 'pressure_veins'] = self.p[4]
        self.data.at[0, 'pressure_ventricle'] = self.p[5]

        self.data.at[0, 'volume_aorta'] = self.v[0]
        self.data.at[0, 'volume_arteries'] = self.v[1]
        self.data.at[0, 'volume_arterioles'] = self.v[2]
        self.data.at[0, 'volume_capillaries'] = self.v[3]
        self.data.at[0, 'volume_veins'] = self.v[4]
        self.data.at[0, 'volume_ventricle'] = self.v[5]

        self.data.at[0, 'ventricle_wall_volume'] = self.ventricle_wall_volume

    def return_flows(self, v):
        # returns fluxes between different compartments

        # Calculate pressure in each compartment
        p = np.zeros(self.no_of_compartments)
        vi = range(self.no_of_compartments-1)
        for x in vi:
            p[x] = v[x] / self.compliance[x]
        p[-1] = self.return_lv_pressure(v[-1])

        flows = dict()

        flows['ventricle_to_aorta'] = 0.0
        if (p[-1] > p[0]):
            flows['ventricle_to_aorta'] = \
                (p[-1] - p[0]) / self.resistance[0]

        flows['aorta_to_arteries'] = \
            (p[0] - p[1]) / self.resistance[1]

        flows['arteries_to_arterioles'] = \
            (p[1] - p[2]) / self.resistance[2]

        flows['arterioles_to_capillaries'] = \
            (p[2] - p[3]) / self.resistance[3]

        flows['capillaries_to_veins'] = \
            (p[3] - p[4]) / self.resistance[4]

        flows['veins_to_ventricle'] = 0.0
        if (p[4] > p[-1]):
            flows['veins_to_ventricle'] = \
                (p[4] - p[-1]) / self.resistance[-1]

        return flows

    def derivs(self, t, v):
        # returns dv, derivative of volume
        dv = np.zeros(self.no_of_compartments)

        # First deduce flows
        flows = self.return_flows(v)

        # Different compartments
        dv[0] = flows['ventricle_to_aorta'] - \
                flows['aorta_to_arteries']

        dv[1] = flows['aorta_to_arteries'] - \
                flows['arteries_to_arterioles']

        dv[2] = flows['arteries_to_arterioles'] - \
                flows['arterioles_to_capillaries']

        dv[3] = flows['arterioles_to_capillaries'] - \
                flows['capillaries_to_veins']

        dv[4] = flows['capillaries_to_veins'] - \
                flows['veins_to_ventricle']

        dv[-1] = flows['veins_to_ventricle'] - \
                flows['ventricle_to_aorta']

        return dv

    def implement_time_step(self, time_step, activation):
        """ Steps circulatory system forward in time """

        # Update the half-sarcomere
        self.hs.update_simulation(time_step, 0.0, activation, 1)

        # steps solution forward in time
        sol = scipy.integrate.solve_ivp(self.derivs, [0, time_step], self.v)
        self.v = sol.y[:, -1]

        # Implements the length change on the half-sarcomere
        new_lv_circumference = self.return_lv_circumference(self.v[-1])
        delta_hsl = self.hs.hs_length *\
            ((new_lv_circumference / self.lv_circumference) - 1.0)
        self.hs.update_simulation(0.0, delta_hsl, 0.0, 1)
        self.lv_circumference = new_lv_circumference

        # Update the pressures
        vi = range(self.no_of_compartments-1)
        for x in vi:
            self.p[x] = self.v[x] / self.compliance[x]
        self.p[-1] = self.return_lv_pressure(self.v[-1])
#        print(self.p)

    def update_data_holders(self, time_step, activation):

        # Update data structure for circulation
        self.sim_time = self.sim_time + time_step
        self.data_buffer_index = self.data_buffer_index + 1
        self.data.at[self.data_buffer_index, 'time'] = self.sim_time
        self.data.at[self.data_buffer_index, 'pressure_aorta'] = self.p[0]
        self.data.at[self.data_buffer_index, 'pressure_arteries'] = self.p[1]
        self.data.at[self.data_buffer_index, 'pressure_arterioles'] = self.p[2]
        self.data.at[self.data_buffer_index, 'pressure_capillaries'] = self.p[3]
        self.data.at[self.data_buffer_index, 'pressure_veins'] = self.p[4]
        self.data.at[self.data_buffer_index, 'pressure_ventricle'] = self.p[-1]
        self.data.at[self.data_buffer_index, 'volume_aorta'] = self.v[0]
        self.data.at[self.data_buffer_index, 'volume_arteries'] = self.v[1]
        self.data.at[self.data_buffer_index, 'volume_arterioles'] = self.v[2]
        self.data.at[self.data_buffer_index, 'volume_capillaries'] = self.v[3]
        self.data.at[self.data_buffer_index, 'volume_veins'] = self.v[4]
        self.data.at[self.data_buffer_index, 'volume_ventricle'] = self.v[-1]

        flows = self.return_flows(self.v)
        self.data.at[self.data_buffer_index, 'flow_ventricle_to_aorta'] = \
            flows['ventricle_to_aorta']
        self.data.at[self.data_buffer_index, 'flow_aorta_to_arteries'] = \
            flows['aorta_to_arteries']
        self.data.at[self.data_buffer_index, 'flow_arteries_to_arterioles'] = \
            flows['arteries_to_arterioles']
        self.data.at[self.data_buffer_index,'flow_arterioles_to_capilllaries'] = \
            flows['arterioles_to_capillaries']
        self.data.at[self.data_buffer_index, 'flow_capillaries_to_veins'] = \
            flows['capillaries_to_veins']
        self.data.at[self.data_buffer_index, 'flow_veins_to_ventricle'] = \
            flows['veins_to_ventricle']

        self.data.at[self.data_buffer_index, 'volume_perturbation'] = \
            self.volume_perturbation[self.data_buffer_index]

        self.data.at[self.data_buffer_index, 'ventricle_wall_volume'] = \
            self.ventricle_wall_volume

        # Now update data structure for half_sarcomere
        self.hs.update_data_holder(time_step, activation)

    def return_lv_circumference(self, lv_volume):
        # 0.001 below is to do with liters to meters conversion
        if (lv_volume > 0.0):
            lv_circum = (2.0 * np.pi * 
                np.power((3 * 0.001 * 
                         (lv_volume + (self.ventricle_wall_volume / 2.0)) /
                         (2 * np.pi)) , (1.0 / 3.0)))
        else:
            lv_circum = (2.0 * np.pi * 
                np.power((3 * 0.001 * 
                         ((self.ventricle_wall_volume / 2.0)) /
                         (2 * np.pi)) , (1.0 / 3.0)))
#        print("lv %f vwv %f" % (lv_volume,self.ventricle_wall_volume))
        return lv_circum

    def return_lv_pressure(self, lv_volume):
        # Deduce new lv circumference
        new_lv_circumference = self.return_lv_circumference(lv_volume)

        # Deduce relative change in hsl
        delta_hsl = self.hs.hs_length * \
            ((new_lv_circumference / self.lv_circumference) - 1.0)

        # Estimate the force produced at the new length
        f = self.hs.myof.check_myofilament_forces(delta_hsl)
        total_force = f['total_force']

#        # Laplaces law says that for a sphere,
#        # P = 2 * S * w / r, where S is wall stress,
#        # w is thickness, and r is internal radius
#        r = np.power((3.0 * 0.001 * lv_volume / (2.0 * np.pi)),(1.0 / 3.0))
#        w = 0.01
#        P_in_pascals = 2 * total_force * w / r
#        P_in_mmHg = P_in_pascals / mmHg_in_pascals
        # Deduce internal radius
        internal_r = np.power((3.0 * 0.001 * lv_volume) /
                              (2.0 * np.pi), (1.0 / 3.0))
        internal_area = 2.0 * np.pi * np.power(internal_r, 2.0)
        wall_thickness = 0.001 * self.ventricle_wall_volume / internal_area
        
        P_in_pascals = 2.0 * total_force * wall_thickness / internal_r
        P_in_mmHg = P_in_pascals / mmHg_in_pascals
        
        return P_in_mmHg
        
        
#        # This equation comes from Slinker and Campbell
#        if (lv_volume>0):
#            return (total_force / mmHg_in_pascals) * (-1 +
#                   np.power((1.0 + (self.ventricle_wall_volume / lv_volume)),(2/3)))
#        else:
#            return 0.0


    def run_simulation(self):
        # Run the simulation
        from .display import display_simulation, display_flows, display_pv_loop

        # Create an activation profile
        no_of_time_points = \
            int(self.simulation_parameters.no_of_time_points.cdata)
        dt = float(self.simulation_parameters.time_step.cdata)
        activation_frequency = \
            float(self.simulation_parameters.activation_frequency.cdata)
        activation_duty_ratio = \
            float(self.simulation_parameters.duty_ratio.cdata)

        # Create an activation pattern
        t = dt*np.arange(1, no_of_time_points+1)
        act = 0.5*(1+signal.square(np.pi+2*np.pi*activation_frequency*t,
                            duty=activation_duty_ratio))

        # Run the simulation
        for i in np.arange(np.size(t)):
            # Apply volume perturbation to veins
            self.v[-2] = self.v[-2] + self.volume_perturbation[i]
            # Update display
            if (np.mod(i, 200)==0):
                print("Blood volume: %.3g, %.0f %% complete" %
                      (np.sum(self.v), (100*i/np.size(t))))
            # Update simulation
            self.implement_time_step(dt, act[i])
            self.update_data_holders(dt, act[i])

        # Concatenate data structures
        self.data = pd.concat([self.data, self.hs.hs_data], axis=1)

        # Make plots
        # Circulation
        display_simulation(self.data,
                           self.output_parameters.summary_figure.cdata)
        display_flows(self.data,
                      self.output_parameters.flows_figure.cdata)
        display_pv_loop(self.data,
                        self.output_parameters.pv_figure.cdata)

        # Half-sarcomere
        hs.half_sarcomere.display_fluxes(self.data,
                               self.output_parameters.hs_fluxes_figure.cdata)

        # Write data to disk
        # Read xml input as a string
        wb = Workbook()
        ws_parameters = wb.active
        ws_parameters.title = 'Simulation parameters'

        tree = etree.parse(self.input_xml_file_string)
        root = tree.getroot()

        def build_xml_string(input_object, current_string, indent):
            
            def indent_string(indent):
                ind_string = ""
                for i in np.arange(0,indent):
                    ind_string = ("%s    " % ind_string)
                return ind_string
            
            for child in input_object:
                current_string = ("%s\n%s<%s>" %
                                  (current_string, indent_string(indent), child.tag))
                if (len(list(child))>0):
                    current_string = build_xml_string(child, current_string, indent+1)
                else:
                    current_string = ("%s%s%s" %
                                     (current_string, indent_string(0), child.text))
                if (len(list(child))==0):
                    current_string = ("%s%s</%s>" %
                                      (current_string, indent_string(0), child.tag))
                else:
                    current_string = ("%s\n%s</%s>" %
                                      (current_string, indent_string(indent), child.tag))
            return current_string

        xml_string = build_xml_string(root,"",0)

        
        if (self.input_xml_file_string):
            f = open(self.input_xml_file_string, mode='r')
            input_xml = f.read()
            f.close
            ws_parameters['A1'] = input_xml
        wb.save(self.output_parameters.data_file.cdata)

        # Append data as a new sheet
        append_df_to_excel(self.output_parameters.data_file.cdata,self.data,
                           sheet_name='Data')

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
